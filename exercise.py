from openpyxl import load_workbook
from openpyxl.styles import PatternFill,Border,Side,Alignment, alignment
import warnings
warnings.filterwarnings('ignore')
fil=load_workbook('./scores.xlsx')
file=fil['重点建设学科专项']
pat=PatternFill('solid',fgColor='EE9572')
side=Side('double')
border=Border(left=side)
align=Alignment(horizontal='left',vertical='center')
file.column_dimensions['D'].width=40
for i in file.iter_rows(min_row=2,max_row=50,min_col=1,max_col=10,values_only=False):
    if i[-4].value>90:
        i[-6].fill=pat
    if i[-2].value=='直博' or i[-2].value=='优博':
        i[-7].border=border
        i[-6].alignment=align
fil.save('./scores.xlsx')
    
        
    
        

