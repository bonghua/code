from openpyxl import load_workbook
import warnings

from openpyxl.workbook.workbook import Workbook
warnings.filterwarnings('ignore')
file=load_workbook(r'./scores.xlsx')
active=file.active
fil=Workbook()
file2=fil.active
dic={}
#for i in active['F']:
#    print(i.value)
for row in active.iter_rows(min_row=2,max_row=500,min_col=1,max_col=10,values_only=False):
    if row[7].value>90 and row[-3].value>80:
        print(row[-5].value)
        ro=[row[-5].value,row[-6].value,row[2].value]
        avg=(row[7].value+row[-3].value)/2
        file2.append(ro)
        if avg>85:
            dic[row[-5].value]={'平均成绩':avg,'院系':row[2].value}
fil.save('90.xlsx')
print(dic)        


   
    